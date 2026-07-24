import json
import base64
from datetime import date, datetime
from io import BytesIO

import requests
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

# ============================================================================
# AYARLAR
# ============================================================================
# NOT (varsayım): Yüklenecek 3 sabit belge aşağıdaki gibi kabul edilmiştir.
# Farklıysa SADECE bu listeyi değiştirmeniz yeterlidir, başka yer değişmez.
REQUIRED_DOCUMENTS = [
    {"key": "ogrenci_belgesi", "label": "Öğrenci Belgesi"},
    {"key": "kimlik_fotokopisi", "label": "Nüfus Cüzdanı / Kimlik Fotokopisi"},
    {"key": "gelir_belgesi", "label": "Aile Gelir Durumunu Gösterir Belge"},
]

MESLEK_SECENEKLERI = [
    "Memur", "İşçi (Özel Sektör)", "Esnaf / Küçük İşletme Sahibi", "Çiftçi / Hayvancılık",
    "Serbest Meslek (Avukat, Doktor, Mühendis vb.)", "Şoför", "Emekli",
    "Ev Hanımı", "İşsiz / Çalışmıyor", "Vefat Etti", "Diğer",
]

TURKCE_AYLAR = [
    "Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
    "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık",
]

try:
    st.set_page_config(page_title="Öğrenci Başvuru Formu", page_icon="logo.png", layout="wide")
except Exception:
    st.set_page_config(page_title="Öğrenci Başvuru Formu", page_icon="📋", layout="wide")

# ============================================================================
# GİRİŞ ALANLARI RENK AYARI (metin kutusu / seçim kutusu arka planı ve yazı rengi)
# ============================================================================
# Sadece bu 2 değeri değiştirerek TÜM metin kutusu, seçim kutusu (selectbox/
# selectbox içindeki "İl seçiniz" gibi yer tutucu yazılar dahil) ve tarih
# seçim kutularının arka plan / yazı rengini tek yerden kontrol edebilirsiniz.
GIRIS_ALANI_ARKA_PLAN_RENGI = "#FFFFFF"  # kutuların arka plan rengi
GIRIS_ALANI_YAZI_RENGI = "#000000"       # kutu içindeki yazı ve yer tutucu rengi (normal siyah)

st.markdown(
    f"""
    <style>
    div[data-baseweb="select"] > div,
    div[data-baseweb="input"],
    input, textarea {{
        background-color: {GIRIS_ALANI_ARKA_PLAN_RENGI} !important;
        color: {GIRIS_ALANI_YAZI_RENGI} !important;
    }}
    div[data-baseweb="select"] span,
    div[data-baseweb="select"] div,
    div[data-baseweb="select"] [class*="placeholder"],
    input::placeholder, textarea::placeholder {{
        color: {GIRIS_ALANI_YAZI_RENGI} !important;
        -webkit-text-fill-color: {GIRIS_ALANI_YAZI_RENGI} !important;
        opacity: 1 !important;
    }}
    </style>
    """,
    unsafe_allow_html=True,
)

@st.cache_data(show_spinner=False)
def _il_ilce_yukle():
    with open("data/il_ilce.json", "r", encoding="utf-8") as f:
        return json.load(f)


@st.cache_data(show_spinner=False)
def _universiteler_yukle():
    with open("data/universiteler.json", "r", encoding="utf-8") as f:
        return json.load(f)


IL_ILCE = _il_ilce_yukle()
IL_LISTESI = list(IL_ILCE.keys())

UNIVERSITE_LISTESI = _universiteler_yukle()
UNIVERSITE_LISTESI_ARAMA = UNIVERSITE_LISTESI + ["Listede Yok / Diğer"]


def _call_apps_script(payload: dict) -> dict:
    """Google Apps Script Web App'e POST isteği gönderir."""
    url = st.secrets["APPS_SCRIPT_URL"]
    payload["secret"] = st.secrets["APPS_SCRIPT_SECRET"]
    resp = requests.post(url, json=payload, timeout=60)
    resp.raise_for_status()
    return resp.json()


def _update_review_status(tc_no, value: str) -> bool:
    payload = {
        "action": "update_review",
        "tc_no": str(tc_no),
        "value": value,
        "secret": st.secrets["APPS_SCRIPT_SECRET"],
    }
    resp = requests.post(st.secrets["APPS_SCRIPT_URL"], json=payload, timeout=30)
    resp.raise_for_status()
    return resp.json().get("ok", False)


@st.cache_data(ttl=15, show_spinner=False)
def _fetch_all_rows() -> dict:
    url = st.secrets["APPS_SCRIPT_URL"]
    resp = requests.get(
        url,
        params={"action": "list_rows", "secret": st.secrets["APPS_SCRIPT_SECRET"]},
        timeout=60,
    )
    resp.raise_for_status()
    return resp.json()


@st.cache_data(ttl=30, show_spinner=False)
def _fetch_settings() -> dict:
    url = st.secrets["APPS_SCRIPT_URL"]
    resp = requests.get(
        url,
        params={"action": "get_settings", "secret": st.secrets["APPS_SCRIPT_SECRET"]},
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json()


@st.cache_data(ttl=30, show_spinner=False)
def _fetch_form_tipi() -> dict:
    """Hangi form tipinin (Tam Anket / Kısa Form) aktif olduğunu getirir (önbelleğe alınır)."""
    url = st.secrets["APPS_SCRIPT_URL"]
    resp = requests.get(
        url,
        params={"action": "get_form_tipi", "secret": st.secrets["APPS_SCRIPT_SECRET"]},
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json()


def _update_form_tipi(form_tipi: str) -> bool:
    payload = {
        "action": "update_form_tipi",
        "form_tipi": form_tipi,
        "secret": st.secrets["APPS_SCRIPT_SECRET"],
    }
    resp = requests.post(st.secrets["APPS_SCRIPT_URL"], json=payload, timeout=30)
    resp.raise_for_status()
    return resp.json().get("ok", False)


@st.cache_data(ttl=15, show_spinner=False)
def _fetch_donemler() -> dict:
    url = st.secrets["APPS_SCRIPT_URL"]
    resp = requests.get(
        url,
        params={"action": "get_donemler", "secret": st.secrets["APPS_SCRIPT_SECRET"]},
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json()


def _sync_donemler(donemler: list) -> dict:
    payload = {
        "action": "sync_donemler",
        "donemler": donemler,
        "secret": st.secrets["APPS_SCRIPT_SECRET"],
    }
    resp = requests.post(st.secrets["APPS_SCRIPT_URL"], json=payload, timeout=30)
    resp.raise_for_status()
    return resp.json()


def _delete_row(tc_no, donem) -> bool:
    payload = {
        "action": "delete_row",
        "tc_no": str(tc_no),
        "donem": str(donem),
        "secret": st.secrets["APPS_SCRIPT_SECRET"],
    }
    resp = requests.post(st.secrets["APPS_SCRIPT_URL"], json=payload, timeout=30)
    resp.raise_for_status()
    return resp.json().get("ok", False)


@st.cache_data(ttl=15, show_spinner=False)
def _fetch_sorular() -> dict:
    """Yönetici paneli için TÜM ek soruları (aktif+pasif) getirir."""
    url = st.secrets["APPS_SCRIPT_URL"]
    resp = requests.get(
        url,
        params={"action": "get_sorular", "secret": st.secrets["APPS_SCRIPT_SECRET"]},
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json()


@st.cache_data(ttl=30, show_spinner=False)
def _fetch_sorular_aktif() -> dict:
    """Form için SADECE aktif ek soruları getirir (önbelleğe alınır)."""
    url = st.secrets["APPS_SCRIPT_URL"]
    resp = requests.get(
        url,
        params={"action": "get_sorular_aktif", "secret": st.secrets["APPS_SCRIPT_SECRET"]},
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json()


def _sync_sorular(sorular: list) -> dict:
    payload = {
        "action": "sync_sorular",
        "sorular": sorular,
        "secret": st.secrets["APPS_SCRIPT_SECRET"],
    }
    resp = requests.post(st.secrets["APPS_SCRIPT_URL"], json=payload, timeout=30)
    resp.raise_for_status()
    return resp.json()


def _file_to_b64(uploaded_file) -> dict:
    data = uploaded_file.getvalue()
    return {
        "filename": uploaded_file.name,
        "mimetype": uploaded_file.type or "application/octet-stream",
        "data_b64": base64.b64encode(data).decode("utf-8"),
    }


def _il_ilce_secimi(il_label: str, ilce_label: str, prefix: str):
    """İl seçilince ilçe listesini otomatik filtreler (canlı güncelleme için form dışında kullanılmalı)."""
    il = st.selectbox(il_label, IL_LISTESI, index=None, placeholder="İl seçiniz", key=f"{prefix}_il")
    ilceler = IL_ILCE.get(il, []) if il else []
    ilce = st.selectbox(
        ilce_label, ilceler, index=None,
        placeholder="Önce il seçiniz" if not il else "İlçe seçiniz",
        disabled=not il, key=f"{prefix}_ilce",
    )
    return il, ilce


def _meslek_secimi(label: str, prefix: str):
    meslek = st.selectbox(label, MESLEK_SECENEKLERI, index=None, placeholder="Seçiniz", key=f"{prefix}_meslek")
    meslek_diger = ""
    if meslek == "Diğer":
        meslek_diger = st.text_input("Mesleği yazın", key=f"{prefix}_meslek_diger")
    return meslek_diger.strip() if meslek == "Diğer" else meslek


@st.cache_data(show_spinner=False)
def _grafikleri_olustur(df: pd.DataFrame, donem_sayilari_df=None):
    """Tüm Genel Bakış grafiklerini bir kez hesaplayıp önbelleğe alır.
    Veri değişmediği sürece (15 sn önbellek süresi içinde) tekrar tıklamalarda
    Plotly grafikleri sıfırdan yeniden hesaplanmaz — panelin kasmasının ana
    sebeplerinden biri buydu."""
    grafikler = {}

    if "Cinsiyet" in df.columns:
        fig = px.pie(df, names="Cinsiyet", title="Cinsiyet Dağılımı", hole=0.5,
                     color_discrete_sequence=["#F5821F", "#2C2C2A"])
        fig.update_layout(margin=dict(t=40, b=0, l=0, r=0))
        grafikler["cinsiyet"] = fig

    if "Sınıfı" in df.columns:
        fig = px.bar(df["Sınıfı"].value_counts().reset_index(), x="Sınıfı", y="count",
                     title="Sınıf Dağılımı", color_discrete_sequence=["#F5821F"])
        fig.update_layout(margin=dict(t=40, b=0, l=0, r=0), xaxis_title="", yaxis_title="")
        grafikler["sinif"] = fig

    if "Bölüm" in df.columns:
        top_bolum = df["Bölüm"].value_counts().nlargest(10).reset_index()
        fig = px.bar(top_bolum, x="count", y="Bölüm", orientation="h",
                     title="En Çok Başvurulan 10 Bölüm", color_discrete_sequence=["#F5821F"])
        fig.update_layout(margin=dict(t=40, b=0, l=0, r=0), xaxis_title="", yaxis_title="",
                           yaxis=dict(autorange="reversed"))
        grafikler["bolum"] = fig

    if "İkamet İl" in df.columns:
        top_il = df["İkamet İl"].value_counts().nlargest(10).reset_index()
        fig = px.bar(top_il, x="count", y="İkamet İl", orientation="h",
                     title="En Çok Başvuru Yapılan 10 İl", color_discrete_sequence=["#2C2C2A"])
        fig.update_layout(margin=dict(t=40, b=0, l=0, r=0), xaxis_title="", yaxis_title="",
                           yaxis=dict(autorange="reversed"))
        grafikler["il"] = fig

    if "Ebeveyn Medeni Durumu" in df.columns:
        fig = px.bar(df["Ebeveyn Medeni Durumu"].value_counts().reset_index(),
                     x="Ebeveyn Medeni Durumu", y="count", title="Ebeveyn Medeni Durumu Dağılımı",
                     color_discrete_sequence=["#F5821F"])
        fig.update_layout(margin=dict(t=40, b=0, l=0, r=0), xaxis_title="", yaxis_title="")
        grafikler["medeni_durum"] = fig

    if "Babasının Mesleği" in df.columns:
        fig = px.bar(df["Babasının Mesleği"].value_counts().nlargest(8).reset_index(),
                     x="count", y="Babasının Mesleği", orientation="h",
                     title="Babasının Mesleği Dağılımı", color_discrete_sequence=["#2C2C2A"])
        fig.update_layout(margin=dict(t=40, b=0, l=0, r=0), xaxis_title="", yaxis_title="",
                           yaxis=dict(autorange="reversed"))
        grafikler["meslek"] = fig

    gerekli_kolonlar = {"Cinsiyet", "Babanın Aylık Geliri (TL)", "Annenin Aylık Geliri (TL)",
                        "Okumakta Olan Kardeş Sayısı"}
    if gerekli_kolonlar.issubset(df.columns) and df["Cinsiyet"].nunique() > 0:
        eksenler = ["Ort. Baba Geliri", "Ort. Anne Geliri", "Ort. Kardeş Sayısı", "Başvuru Oranı (%)"]
        fig = go.Figure()
        renkler = {"Kadın": "#F5821F", "Erkek": "#2C2C2A"}
        for cinsiyet_grubu, renk in renkler.items():
            alt_df = df[df["Cinsiyet"] == cinsiyet_grubu]
            if alt_df.empty:
                continue
            degerler = [
                alt_df["Babanın Aylık Geliri (TL)"].fillna(0).mean(),
                alt_df["Annenin Aylık Geliri (TL)"].fillna(0).mean(),
                alt_df["Okumakta Olan Kardeş Sayısı"].fillna(0).mean() * 5000,
                (len(alt_df) / len(df)) * 100 * 200,
            ]
            fig.add_trace(go.Scatterpolar(
                r=degerler, theta=eksenler, fill="toself", name=cinsiyet_grubu, line_color=renk,
            ))
        fig.update_layout(
            title="Cinsiyete Göre Profil Karşılaştırması (Örümcek Grafik)",
            polar=dict(radialaxis=dict(visible=False)),
            margin=dict(t=40, b=0, l=40, r=40), showlegend=True,
        )
        grafikler["radar"] = fig

    if "Babanın Aylık Geliri (TL)" in df.columns and "Annenin Aylık Geliri (TL)" in df.columns:
        toplam_gelir = df["Babanın Aylık Geliri (TL)"].fillna(0) + df["Annenin Aylık Geliri (TL)"].fillna(0)
        fig = px.histogram(toplam_gelir, nbins=15, title="Aile Toplam Gelir Dağılımı",
                           color_discrete_sequence=["#F5821F"])
        fig.update_layout(margin=dict(t=40, b=0, l=0, r=0), xaxis_title="Toplam Gelir (TL)",
                           yaxis_title="Başvuru Sayısı", showlegend=False)
        grafikler["gelir_histogram"] = fig

    if donem_sayilari_df is not None and len(donem_sayilari_df) > 1:
        fig = px.bar(donem_sayilari_df, x="Dönem", y="count", title="Dönemlere Göre Toplam Başvuru Sayısı",
                     color_discrete_sequence=["#F5821F"])
        fig.update_layout(margin=dict(t=40, b=0, l=0, r=0), xaxis_title="", yaxis_title="")
        grafikler["donem_karsilastirma"] = fig

    return grafikler


def _turkce_tarih_secimi(label: str, prefix: str, varsayilan: date, min_yil: int, max_yil: int):
    """Takvim yerine Gün/Ay/Yıl seçim kutuları — ay isimleri Türkçe, tam kontrol bizde."""
    st.markdown(f"<p style='font-size:14px; color:#31333F; margin-bottom:2px;'>{label}</p>", unsafe_allow_html=True)
    c1, c2, c3 = st.columns(3)
    with c1:
        gun = st.selectbox(
            "Gün", list(range(1, 32)), index=varsayilan.day - 1,
            key=f"{prefix}_gun", label_visibility="collapsed",
        )
    with c2:
        ay = st.selectbox(
            "Ay", TURKCE_AYLAR, index=varsayilan.month - 1,
            key=f"{prefix}_ay", label_visibility="collapsed",
        )
    with c3:
        yil_listesi = list(range(max_yil, min_yil - 1, -1))
        yil = st.selectbox(
            "Yıl", yil_listesi, index=yil_listesi.index(varsayilan.year),
            key=f"{prefix}_yil", label_visibility="collapsed",
        )
    ay_no = TURKCE_AYLAR.index(ay) + 1
    try:
        return date(yil, ay_no, gun)
    except ValueError:
        return None  # örn. 30 Şubat gibi geçersiz kombinasyon


def _logo_base64():
    try:
        with open("logo.png", "rb") as f:
            return base64.b64encode(f.read()).decode("utf-8")
    except Exception:
        return None


def _ust_serit(baslik: str, alt_baslik: str, buton_metni: str = None, buton_key: str = "ust_serit_btn") -> bool:
    logo_b64 = _logo_base64()
    st.markdown(
        """
        <style>
        .st-key-ust_serit_kutu { background:#F5821F; border-radius:12px;
            padding:1.25rem 2rem; margin-bottom:1.5rem; }
        .st-key-ust_serit_kutu button { background:#FFFFFF !important; }
        </style>
        """,
        unsafe_allow_html=True,
    )
    tiklandi = False
    # NOT: Sol ve sağ kenar kolonları BİLEREK eşit toplam genişlikte (3 birim) tutulur.
    # Bu sayede logo kolonu, içeriğin uzunluğundan bağımsız olarak matematiksel
    # olarak tam ortada kalır (position:absolute yerine simetrik kolon genişliği kullanıldı,
    # çünkü absolute konumlandırma bazı Streamlit sürümlerinde şeridin dışına taşabiliyor).
    with st.container(key="ust_serit_kutu"):
        col_baslik, col_logo, col_bosluk, col_buton = st.columns(
            [3, 1.4, 1.4, 1.6], vertical_alignment="center"
        )
        with col_baslik:
            st.markdown(
                f"<p style='color:#FFFFFF; font-size:24px; font-weight:600; margin:0;'>{baslik}</p>"
                f"<p style='color:#FFFFFF; font-size:14px; margin:6px 0 0; opacity:0.92;'>{alt_baslik}</p>",
                unsafe_allow_html=True,
            )
        with col_logo:
            if logo_b64:
                # LOGO_BOYUT: logo çapı (px). LOGO_DIKEY_KAYDIRMA: negatif değer yukarı,
                # pozitif değer aşağı kaydırır — sadece bu iki değeri değiştirerek ayarlayın.
                LOGO_BOYUT = 100
                LOGO_DIKEY_KAYDIRMA = "-10px"
                st.markdown(
                    f'<div style="text-align:center;">'
                    f'<img src="data:image/png;base64,{logo_b64}" '
                    f'style="height:{LOGO_BOYUT}px; width:{LOGO_BOYUT}px; border-radius:50%; '
                    f'background:#FFFFFF; padding:4px; object-fit:cover; display:inline-block; '
                    f'vertical-align:middle; position:relative; top:{LOGO_DIKEY_KAYDIRMA};" /></div>',
                    unsafe_allow_html=True,
                )
        with col_buton:
            if buton_metni:
                tiklandi = st.button(buton_metni, key=buton_key, use_container_width=True)
    return tiklandi


def _bolum_basligi(numara: str, baslik: str):
    st.markdown(
        f"""
        <div style="display:flex; align-items:center; gap:10px; margin:1.75rem 0 1rem;">
            <div style="width:26px; height:26px; border-radius:50%; background:#F5821F;
                        color:#FFFFFF; font-size:13px; font-weight:600; display:flex;
                        align-items:center; justify-content:center; flex-shrink:0;">{numara}</div>
            <p style="font-size:16px; font-weight:600; margin:0; color:#111111;">{baslik}</p>
        </div>
        <hr style="margin:-6px 0 1rem; border:none; border-top:1px solid #E5E5E5;">
        """,
        unsafe_allow_html=True,
    )


@st.dialog("Silme İşlemini Onaylayın")
def _silme_onay_dialogu():
    n = len(st.session_state.get("silinecek_kayitlar", set()))
    st.write(f"**{n} adet kayıt silinecektir.**")
    st.write(
        "Bu işlem sonrasında başvuru satırı tablodan kalıcı olarak silinir "
        "(Drive'daki evrak klasörü çöp kutusuna taşınır, 30 gün içinde geri alınabilir)."
    )
    st.write("**Emin misiniz?**")
    col1, col2 = st.columns(2)
    with col1:
        if st.button("Evet, Sil", type="primary", use_container_width=True):
            basarili = 0
            for tc_no, donem_deger in list(st.session_state.silinecek_kayitlar):
                try:
                    if _delete_row(tc_no, donem_deger):
                        basarili += 1
                except Exception as e:
                    st.error(f"{tc_no} silinemedi: {e}")
            st.session_state.silinecek_kayitlar = set()
            st.session_state["silme_onay_bekliyor"] = False
            _fetch_all_rows.clear()
            st.session_state["silme_basarili_sayisi"] = basarili
            st.rerun()
    with col2:
        if st.button("Vazgeç", use_container_width=True):
            st.session_state["silme_onay_bekliyor"] = False
            st.rerun()


@st.dialog("Başvuruyu Onaylayın")
def _onay_dialogu():
    st.write("Bilgilerinizi kontrol ettiniz mi? Başvuru gönderildikten sonra düzenleyemezsiniz.")
    st.write("**Başvuruyu göndermek istediğinizden emin misiniz?**")
    col1, col2 = st.columns(2)
    with col1:
        if st.button("Evet, Gönder", type="primary", use_container_width=True):
            with st.spinner("Başvurunuz gönderiliyor, lütfen bekleyin..."):
                try:
                    result = _call_apps_script(st.session_state["pending_payload"])
                    if result.get("ok"):
                        st.session_state["confirm_pending"] = False
                        st.session_state["submit_success"] = True
                        st.rerun()
                    else:
                        st.error(f"Gönderim başarısız: {result.get('error', 'Bilinmeyen hata')}")
                except Exception as e:
                    st.error(f"Bir hata oluştu, lütfen tekrar deneyin. Teknik detay: {e}")
    with col2:
        if st.button("Vazgeç", use_container_width=True):
            st.session_state["confirm_pending"] = False
            st.rerun()


# ============================================================================
# NAVİGASYON
# ============================================================================
if "view" not in st.session_state:
    st.session_state.view = "form"
if "admin_authed" not in st.session_state:
    st.session_state.admin_authed = False

# ============================================================================
# ÖĞRENCİ BAŞVURU FORMU
# ============================================================================
# NOT: Bu bölüm bilinçli olarak st.form KULLANMAZ. İl -> İlçe seçimlerinin
# anında (her il değiştiğinde) filtrelenebilmesi için widget'ların her
# etkileşimde yeniden çalışması (rerun) gerekir; st.form içindeki widget'lar
# sadece "gönder" butonuna basılınca güncellenir ve ilçe listesi o ana kadar
# bayat kalırdı. Bu yüzden gönderim ayrı bir st.button ile tetiklenir.
if st.session_state.view == "form":
    try:
        ayarlar = _fetch_settings()
    except Exception as e:
        ayarlar = {"ok": False}
        st.error(f"Form ayarları alınamadı, lütfen sayfayı yenileyin. Teknik detay: {e}")

    donem_adi = ayarlar.get("donem", "") if ayarlar.get("ok") else ""
    aktif_mi = ayarlar.get("aktif") == "Evet" if ayarlar.get("ok") else False

    alt_baslik = f"Bol-Dav Bolvadinliler Dayanışma Vakfı — {donem_adi}" if donem_adi else "Bol-Dav Bolvadinliler Dayanışma Vakfı"
    if _ust_serit("Bol-Dav Öğrenci Burs Bilgi ve Başvuru Formu", alt_baslik, "Yönetici Girişi", "btn_yonetici_girisi"):
        st.session_state.view = "admin"
        st.rerun()

    if not ayarlar.get("ok"):
        st.stop()

    if not aktif_mi:
        st.warning("📌 Başvurular şu anda kapalıdır. Yeni dönem başvuruları açıldığında bu sayfadan duyurulacaktır.")
        st.stop()

    try:
        form_tipi_sonuc = _fetch_form_tipi()
        form_tipi = form_tipi_sonuc.get("form_tipi", "Tam") if form_tipi_sonuc.get("ok") else "Tam"
    except Exception:
        form_tipi = "Tam"
    kisa_form_mu = form_tipi == "Kisa"

    st.write("Lütfen aşağıdaki bilgileri eksiksiz doldurun ve istenen belgeleri ekleyin. Tüm alanlar zorunludur.")

    _bolum_basligi("1", "Öğrenci Bilgileri")
    c1, c2 = st.columns(2)
    with c1:
        ad_soyad = st.text_input("Öğrenci Adı Soyadı *")
        tc_no = st.text_input("T.C. Kimlik No *", max_chars=11)
    with c2:
        cinsiyet = st.radio("Cinsiyet *", ["Erkek", "Kadın"], horizontal=True)
        email = st.text_input("E-Posta Adresi *")
    telefon = st.text_input("Cep Telefon Numarası *", placeholder="05xx xxx xx xx")
    dogum_tarihi = _turkce_tarih_secimi(
        "Öğrenci Doğum Tarihi *", "ogrenci_dogum_tarihi",
        varsayilan=date(2004, 1, 1), min_yil=1980, max_yil=date.today().year,
    )

    st.markdown("**Öğrenci Doğum Yeri**")
    d1, d2 = st.columns(2)
    with d1:
        dogum_il, dogum_ilce = _il_ilce_secimi("İl *", "İlçe *", "ogrenci_dogum")
    st.markdown("**Nüfusa Kayıtlı Olduğu Yer**")
    n1, n2 = st.columns(2)
    with n1:
        nufus_il, nufus_ilce = _il_ilce_secimi("İl *", "İlçe *", "nufus_kayit")

    # -------- Aşağıdaki bölümler SADECE "Tam Anket" form tipinde gösterilir --------
    ikamet_il = ikamet_ilce = kaldigi_yer = None
    universite = universite_diger = bolum = okul_ili = None
    fakulte = sinif = lise_derece = lise_adi = None
    baba_adi = baba_telefon = baba_meslek = baba_adres = None
    baba_gelir = baba_dogum_tarihi = baba_dogum_il = baba_dogum_ilce = None
    anne_adi = anne_telefon = anne_meslek = anne_adres = None
    anne_gelir = anne_dogum_tarihi = anne_dogum_il = anne_dogum_ilce = None
    ebeveyn_durumu = None
    kardes_sayisi = kardes_okullari = sosyo_ekonomik = None
    aktif_sorular = []
    ek_cevaplar = {}

    if not kisa_form_mu:
        _bolum_basligi("2", "İkamet Bilgisi")
        c3, c4 = st.columns(2)
        with c3:
            ikamet_il, ikamet_ilce = _il_ilce_secimi("İkamet Edilen İl *", "İkamet Edilen İlçe *", "ikamet")
        with c4:
            kaldigi_yer = st.selectbox(
                "Öğrencinin Kaldığı Yer *",
                ["Aile Yanında", "Yurtta", "Kirada / Arkadaşlarıyla", "Akraba Yanında", "Diğer"],
            )

        _bolum_basligi("3", "Eğitim Bilgileri")
        c5, c6 = st.columns(2)
        with c5:
            universite = st.selectbox(
                "Üniversite Adı *", UNIVERSITE_LISTESI_ARAMA, index=None,
                placeholder="Üniversite adı yazarak arayın...",
            )
            universite_diger = ""
            if universite == "Listede Yok / Diğer":
                universite_diger = st.text_input("Üniversite adını yazın *")
            bolum = st.text_input("Bölüm *")
            okul_ili = st.selectbox("Okulun Bulunduğu İl *", IL_LISTESI, index=None, placeholder="İl seçiniz")
        with c6:
            fakulte = st.text_input("Fakülte *")
            sinif = st.selectbox("Sınıfı *", ["Hazırlık", "1", "2", "3", "4", "5", "6", "Yüksek Lisans"])
            lise_derece = st.text_input("Lise Mezuniyet Derecesi / Ortalaması *")
        lise_adi = st.text_input("Mezun Olduğu Lise ve Dengi Okul *")

        _bolum_basligi("4", "Aile Bilgileri")
        st.markdown("**Baba**")
        c7, c8 = st.columns(2)
        with c7:
            baba_adi = st.text_input("Baba Adı *")
            baba_telefon = st.text_input("Baba Telefonu *")
        with c8:
            baba_meslek = _meslek_secimi("Babasının Mesleği *", "baba")
            baba_gelir = st.number_input(
                "Babanın Aylık Gelir Durumu (TL) *", min_value=0, step=500,
                value=None, placeholder="Örn: 15000",
            )
        baba_dogum_tarihi = _turkce_tarih_secimi(
            "Baba Doğum Tarihi *", "baba_dogum_tarihi",
            varsayilan=date(1975, 1, 1), min_yil=1930, max_yil=date.today().year,
        )
        st.markdown("Baba Doğum Yeri")
        bd1, bd2 = st.columns(2)
        with bd1:
            baba_dogum_il, baba_dogum_ilce = _il_ilce_secimi("İl *", "İlçe *", "baba_dogum")
        baba_adres = st.text_area("Baba Adresi *")

        st.markdown("**Anne**")
        c9, c10 = st.columns(2)
        with c9:
            anne_adi = st.text_input("Anne Adı *")
            anne_telefon = st.text_input("Anne Telefonu *")
        with c10:
            anne_meslek = _meslek_secimi("Annenin Mesleği *", "anne")
            anne_gelir = st.number_input(
                "Annenin Aylık Gelir Durumu (TL) *", min_value=0, step=500,
                value=None, placeholder="Örn: 10000",
            )
        anne_dogum_tarihi = _turkce_tarih_secimi(
            "Anne Doğum Tarihi *", "anne_dogum_tarihi",
            varsayilan=date(1978, 1, 1), min_yil=1930, max_yil=date.today().year,
        )
        st.markdown("Anne Doğum Yeri")
        ad1, ad2 = st.columns(2)
        with ad1:
            anne_dogum_il, anne_dogum_ilce = _il_ilce_secimi("İl *", "İlçe *", "anne_dogum")
        anne_adres = st.text_area("Anne Adresi *")

        ebeveyn_durumu = st.selectbox(
            "Ebeveyn Medeni Durumu *",
            ["Evli", "Boşanmış", "Baba Vefat", "Anne Vefat", "Her İkisi Vefat"],
        )

        _bolum_basligi("5", "Sosyoekonomik Bilgiler")
        kardes_sayisi = st.number_input("Okumakta Olan Kardeş Sayısı *", min_value=0, max_value=15, step=1)
        kardes_okullari = st.text_area(
            "Kardeşlerin Okuduğu Okullar *",
            placeholder="Okuyan kardeşiniz yoksa 'Yok' yazınız",
        )
        sosyo_ekonomik = st.text_area(
            "Sosyo Ekonomik Faktörler *",
            placeholder="Aile durumunu etkileyen ek faktörler (engellilik, kronik hastalık, kira yükü vb.)",
        )

        try:
            aktif_sorular = _fetch_sorular_aktif().get("sorular", [])
        except Exception:
            aktif_sorular = []

        if aktif_sorular:
            _bolum_basligi("6", "Ek Sorular")
            for soru in aktif_sorular:
                soru_id = soru["id"]
                soru_metni = soru["soru"] + " *"
                tip = soru.get("tip", "Kısa Metin")
                if tip == "Uzun Metin":
                    ek_cevaplar[soru_id] = st.text_area(soru_metni, key=f"eksoru_{soru_id}")
                elif tip == "Sayı":
                    ek_cevaplar[soru_id] = st.number_input(soru_metni, step=1, key=f"eksoru_{soru_id}")
                elif tip == "Tarih":
                    ek_cevaplar[soru_id] = _turkce_tarih_secimi(
                        soru_metni, f"eksoru_{soru_id}",
                        varsayilan=date.today(), min_yil=1930, max_yil=date.today().year + 1,
                    )
                elif tip == "Seçenekli":
                    secenekler = [s.strip() for s in soru.get("secenekler", "").split(",") if s.strip()]
                    ek_cevaplar[soru_id] = st.selectbox(
                        soru_metni, secenekler, index=None, placeholder="Seçiniz", key=f"eksoru_{soru_id}"
                    )
                elif tip == "Evet/Hayır":
                    ek_cevaplar[soru_id] = st.radio(soru_metni, ["Evet", "Hayır"], horizontal=True, key=f"eksoru_{soru_id}")
                else:  # Kısa Metin
                    ek_cevaplar[soru_id] = st.text_input(soru_metni, key=f"eksoru_{soru_id}")

    # -------- Belgeler: SADECE "Kısa Form" tipinde gösterilir --------
    uploaded = {}
    if kisa_form_mu:
        _bolum_basligi("2", "Belgeler")
        st.caption("Kabul edilen dosya türleri: PDF, JPG, PNG — dosya başına en fazla 10 MB.")
        for doc in REQUIRED_DOCUMENTS:
            uploaded[doc["key"]] = st.file_uploader(
                f"{doc['label']} *", type=["pdf", "jpg", "jpeg", "png"], key=doc["key"]
            )

    if kisa_form_mu:
        onay_no = "3"
    else:
        onay_no = "7" if aktif_sorular else "6"
    ONAY_IFADESI = "Bol-Dav Burs Başvurusu Yükümlülük Taahhütnamesini Okudum Onaylıyorum"

    _bolum_basligi(onay_no, "Onay")
    if kisa_form_mu:
        st.markdown(
            """
            <div style="background:#FAFAFA; border:1px solid #E5E5E5; border-radius:8px;
                        padding:1.25rem 1.5rem; font-size:14px; line-height:1.8; color:#111111;">
                <p style="text-align:center; font-weight:600; margin:0 0 12px;">BOL-DAV BURSLARI KARŞILIKSIZDIR</p>
                <p>Öğrenim hayatım boyunca akademik başarı düzeyimi, BOL-DAV Vakfı Burs Verme Usul ve
                Esaslarında belirtilen kriterlerin üzerinde tutmak için azami gayret göstereceğimi;
                vakfımızın teşvik ettiği sosyal, kültürel ve toplumsal sorumluluk projelerine aktif olarak
                katılım sağlamaya çalışacağımı beyan ediyorum.</p>
                <p>Allah nasip eder de eğitimimi tamamlayıp çalışma hayatına atılırsam, imkânlarım
                ölçüsünde BOL-DAV'a maddi ve manevi destek vermeyi; bir zamanlar bana uzanan yardım elini,
                kendim gibi eğitim hayatını sürdüren en az bir Bolvadinli üniversite öğrencisine ulaştırmak
                amacıyla BOL-DAV aracılığıyla burs vermeyi hedef olarak benimsediğimi ifade ediyorum.</p>
                <p>Ayrıca, BOL-DAV'ın amaç ve faaliyetlerini bulunduğum her ortamda tanıtmayı, vakfımızın
                birlik ve dayanışma anlayışını temsil etmeyi, gönüllü bir kültür ve dayanışma elçisi olarak
                hareket etmeyi vicdani bir sorumluluk ve vefa borcu olarak kabul ediyorum.</p>
                <p style="text-align:center; font-weight:600; margin:16px 0 0;">BOL-DAV BOLVADİNLİLER DAYANIŞMA VAKFI</p>
            </div>
            """,
            unsafe_allow_html=True,
        )
        st.write("")
        st.markdown(
            f"Onaylamak için aşağıdaki kutuya **tam olarak** şunu yazın: "
            f"\n\n> {ONAY_IFADESI}"
        )
        onay_metni = st.text_input("Onay ifadesi *", placeholder=ONAY_IFADESI)
        onay = onay_metni.strip() == ONAY_IFADESI
    else:
        onay = st.checkbox("Yukarıdaki bilgileri doğrularım ve belgelendiririm. *")

    submitted = st.button("Başvuruyu Gönder", use_container_width=True, type="primary")

    if submitted:
        zorunlu_metin_alanlari = {
            "Öğrenci Adı Soyadı": ad_soyad, "T.C. Kimlik No": tc_no, "E-Posta Adresi": email,
            "Cep Telefon Numarası": telefon,
        }
        if not kisa_form_mu:
            zorunlu_metin_alanlari.update({
                "Fakülte": fakulte, "Bölüm": bolum, "Mezun Olduğu Lise": lise_adi,
                "Lise Mezuniyet Derecesi": lise_derece,
                "Baba Adı": baba_adi, "Baba Telefonu": baba_telefon, "Baba Adresi": baba_adres,
                "Anne Adı": anne_adi, "Anne Telefonu": anne_telefon, "Anne Adresi": anne_adres,
                "Sosyo Ekonomik Faktörler": sosyo_ekonomik, "Kardeşlerin Okuduğu Okullar": kardes_okullari,
            })
        eksikler = [ad for ad, deger in zorunlu_metin_alanlari.items() if not (deger or "").strip()]

        il_ilce_kontrol_listesi = [
            (dogum_il, dogum_ilce, "Öğrenci Doğum Yeri (İl/İlçe)"),
            (nufus_il, nufus_ilce, "Nüfusa Kayıtlı Olduğu Yer (İl/İlçe)"),
        ]
        if not kisa_form_mu:
            il_ilce_kontrol_listesi += [
                (ikamet_il, ikamet_ilce, "İkamet Edilen İl/İlçe"),
                (baba_dogum_il, baba_dogum_ilce, "Baba Doğum Yeri (İl/İlçe)"),
                (anne_dogum_il, anne_dogum_ilce, "Anne Doğum Yeri (İl/İlçe)"),
            ]
        for il_deger, ilce_deger, ad in il_ilce_kontrol_listesi:
            if not il_deger or not ilce_deger:
                eksikler.append(ad)

        universite_final = None
        if not kisa_form_mu:
            if not okul_ili:
                eksikler.append("Okulun Bulunduğu İl")
            if not universite:
                eksikler.append("Üniversite Adı")
            elif universite == "Listede Yok / Diğer" and not universite_diger.strip():
                eksikler.append("Üniversite Adı (listede yoksa elle yazılmalı)")
            universite_final = universite_diger.strip() if universite == "Listede Yok / Diğer" else universite

            if not baba_meslek:
                eksikler.append("Babasının Mesleği")
            if not anne_meslek:
                eksikler.append("Annenin Mesleği")
            if baba_gelir is None:
                eksikler.append("Babanın Aylık Gelir Durumu")
            if anne_gelir is None:
                eksikler.append("Annenin Aylık Gelir Durumu")

        tarih_kontrol_listesi = [(dogum_tarihi, "Öğrenci Doğum Tarihi (geçersiz gün/ay kombinasyonu)")]
        if not kisa_form_mu:
            tarih_kontrol_listesi += [
                (baba_dogum_tarihi, "Baba Doğum Tarihi (geçersiz gün/ay kombinasyonu)"),
                (anne_dogum_tarihi, "Anne Doğum Tarihi (geçersiz gün/ay kombinasyonu)"),
            ]
        for tarih_deger, ad in tarih_kontrol_listesi:
            if tarih_deger is None:
                eksikler.append(ad)

        for soru in aktif_sorular:
            cevap = ek_cevaplar.get(soru["id"])
            if soru.get("tip") in ("Kısa Metin", "Uzun Metin") and not str(cevap or "").strip():
                eksikler.append(soru["soru"])
            elif soru.get("tip") == "Seçenekli" and not cevap:
                eksikler.append(soru["soru"])
            elif soru.get("tip") == "Tarih" and cevap is None:
                eksikler.append(soru["soru"] + " (geçersiz gün/ay kombinasyonu)")

        if len(tc_no.strip()) != 11 or not tc_no.strip().isdigit():
            eksikler.append("T.C. Kimlik No (11 haneli olmalı)")
        if kisa_form_mu:
            for doc in REQUIRED_DOCUMENTS:
                if uploaded.get(doc["key"]) is None:
                    eksikler.append(doc["label"])
        if not onay:
            if kisa_form_mu:
                eksikler.append("Taahhütname ifadesi tam olarak yazılmalı")
            else:
                eksikler.append("Onay kutusu işaretlenmeli")

        if eksikler:
            st.error("Aşağıdaki alanları tamamlayın:\n\n- " + "\n- ".join(eksikler))
        else:
            temel_alanlar = {
                "Form Tipi": "Kısa Form" if kisa_form_mu else "Tam Anket",
                "Cinsiyet": cinsiyet, "Öğrenci Adı Soyadı": ad_soyad, "T.C. Kimlik No": tc_no,
                "E-Posta Adresi": email, "Cep Telefon Numarası": telefon,
                "Öğrenci Doğum Tarihi": str(dogum_tarihi),
                "Öğrenci Doğum Yeri İl": dogum_il, "Öğrenci Doğum Yeri İlçe": dogum_ilce,
                "Nüfusa Kayıtlı Olduğu İl": nufus_il, "Nüfusa Kayıtlı Olduğu İlçe": nufus_ilce,
            }
            if not kisa_form_mu:
                temel_alanlar.update({
                    "İkamet İl": ikamet_il, "İkamet İlçe": ikamet_ilce,
                    "Öğrencinin Kaldığı Yer": kaldigi_yer,
                    "Üniversite Adı": universite_final, "Fakülte": fakulte, "Bölüm": bolum, "Sınıfı": sinif,
                    "Okulun Bulunduğu İl": okul_ili, "Mezun Olduğu Lise": lise_adi,
                    "Lise Mezuniyet Derecesi": lise_derece,
                    "Baba Adı": baba_adi,
                    "Baba Doğum Yeri İl": baba_dogum_il, "Baba Doğum Yeri İlçe": baba_dogum_ilce,
                    "Baba Doğum Tarihi": str(baba_dogum_tarihi), "Baba Adresi": baba_adres,
                    "Baba Telefonu": baba_telefon, "Babasının Mesleği": baba_meslek,
                    "Babanın Aylık Geliri (TL)": baba_gelir,
                    "Anne Adı": anne_adi,
                    "Anne Doğum Yeri İl": anne_dogum_il, "Anne Doğum Yeri İlçe": anne_dogum_ilce,
                    "Anne Doğum Tarihi": str(anne_dogum_tarihi), "Anne Adresi": anne_adres,
                    "Anne Telefonu": anne_telefon, "Annenin Mesleği": anne_meslek,
                    "Annenin Aylık Geliri (TL)": anne_gelir,
                    "Ebeveyn Medeni Durumu": ebeveyn_durumu,
                    "Okumakta Olan Kardeş Sayısı": kardes_sayisi,
                    "Kardeşlerin Okuduğu Okullar": kardes_okullari,
                    "Sosyo Ekonomik Faktörler": sosyo_ekonomik,
                    **{
                        f"Ek Soru: {soru['soru']}": str(ek_cevaplar.get(soru["id"], ""))
                        for soru in aktif_sorular
                    },
                })

            payload = {
                "action": "submit",
                "timestamp": datetime.now().isoformat(),
                "fields": temel_alanlar,
                "files": (
                    {doc["key"]: _file_to_b64(uploaded[doc["key"]]) for doc in REQUIRED_DOCUMENTS}
                    if kisa_form_mu else {}
                ),
                "folder_name": f"{ad_soyad.strip()}_{tc_no.strip()}",
            }
            st.session_state["pending_payload"] = payload
            st.session_state["confirm_pending"] = True
            st.rerun()

    if st.session_state.get("confirm_pending"):
        _onay_dialogu()

    if st.session_state.get("submit_success"):
        st.success("Başvurunuz başarıyla alındı. Teşekkür ederiz.")
        st.balloons()
        st.session_state["submit_success"] = False


# ============================================================================
# YÖNETİCİ PANELİ
# ============================================================================
else:
    if _ust_serit("Yönetici Paneli", "Bol-Dav Bolvadinliler Dayanışma Vakfı", "Forma Dön", "btn_forma_don"):
        st.session_state.view = "form"
        st.rerun()

    if not st.session_state.admin_authed:
        st.markdown("#### Yönetici Girişi")
        u = st.text_input("Kullanıcı Adı")
        p = st.text_input("Şifre", type="password")
        if st.button("Giriş Yap"):
            creds = st.secrets["admin_credentials"]
            if u == creds["user"] and p == creds["pass"]:
                st.session_state.admin_authed = True
                st.rerun()
            else:
                st.error("Kullanıcı adı veya şifre hatalı.")
    else:
        if st.button("🔄 Verileri Yenile"):
            st.cache_data.clear()

        if "exp_form_tipi_acik" not in st.session_state:
            st.session_state.exp_form_tipi_acik = False
        with st.expander("🧾 Form Tipi (Hangi form öğrenciye gösterilsin?)", expanded=st.session_state.exp_form_tipi_acik):
            st.caption(
                "**Tam Anket:** Mevcut tüm sorular (öğrenci/aile/eğitim/sosyoekonomik bilgiler) gösterilir, "
                "**belge yükleme yoktur**.\n\n"
                "**Kısa Form:** Sadece 1. Öğrenci Bilgileri bölümü + belge yükleme gösterilir, "
                "diğer tüm bölümler (ikamet, eğitim, aile, sosyoekonomik, ek sorular) formda görünmez."
            )
            if st.session_state.get("form_tipi_basarili"):
                st.success("✅ Form tipi güncellendi.")
                st.session_state.form_tipi_basarili = False
            try:
                form_tipi_sonuc = _fetch_form_tipi()
            except Exception as e:
                form_tipi_sonuc = {"ok": False}
                st.error(f"Form tipi alınamadı: {e}")

            if form_tipi_sonuc.get("ok"):
                mevcut_form_tipi = form_tipi_sonuc.get("form_tipi", "Tam")
                secim = st.radio(
                    "Aktif form tipi",
                    ["Tam Anket (Dosyasız)", "Kısa Form (Öğrenci Bilgisi + Belge)"],
                    index=0 if mevcut_form_tipi == "Tam" else 1,
                )
                if st.button("💾 Form Tipini Kaydet", type="primary"):
                    yeni_deger = "Tam" if secim == "Tam Anket (Dosyasız)" else "Kisa"
                    try:
                        if _update_form_tipi(yeni_deger):
                            _fetch_form_tipi.clear()
                            st.session_state.form_tipi_basarili = True
                            st.session_state.exp_form_tipi_acik = True
                            st.rerun()
                        else:
                            st.error("Form tipi güncellenemedi.")
                    except Exception as e:
                        st.error(f"Form tipi güncellenemedi: {e}")

        if "exp_donem_acik" not in st.session_state:
            st.session_state.exp_donem_acik = False
        with st.expander("⚙️ Dönem Yönetimi (Ekle / Düzenle / Sil / Aktif-Pasif)", expanded=st.session_state.exp_donem_acik):
            if st.session_state.get("donem_basarili"):
                st.success("✅ Dönemler güncellendi.")
                st.session_state.donem_basarili = False
            try:
                donem_sonuc = _fetch_donemler()
            except Exception as e:
                donem_sonuc = {"ok": False}
                st.error(f"Dönemler alınamadı: {e}")

            if donem_sonuc.get("ok"):
                st.caption(
                    "Tabloya yeni satır ekleyerek yeni dönem oluşturun, isimleri doğrudan düzenleyin, "
                    "satır silmek için satırın solundaki kutucuğu işaretleyip çöp kutusuna basın. "
                    "**Aynı anda yalnızca bir dönem 'Aktif' olabilir** — o dönem formda görünür ve "
                    "başvuruları/dosyaları kendi Drive klasörüne kaydeder."
                )
                donem_df = pd.DataFrame(donem_sonuc.get("donemler", []))
                if donem_df.empty:
                    donem_df = pd.DataFrame([{"ad": "", "durum": "Pasif"}])

                edited_donem_df = st.data_editor(
                    donem_df,
                    use_container_width=True,
                    hide_index=True,
                    num_rows="dynamic",
                    column_config={
                        "ad": st.column_config.TextColumn("Dönem Adı", required=True),
                        "durum": st.column_config.SelectboxColumn("Durum", options=["Aktif", "Pasif"], required=True),
                    },
                    key="donem_editor",
                )

                if st.button("💾 Dönemleri Kaydet", type="primary"):
                    temiz_liste = [
                        {"ad": str(r["ad"]).strip(), "durum": r["durum"]}
                        for _, r in edited_donem_df.iterrows()
                        if str(r["ad"]).strip()
                    ]
                    aktif_sayisi = sum(1 for d in temiz_liste if d["durum"] == "Aktif")
                    if aktif_sayisi > 1:
                        st.error("Aynı anda sadece bir dönem 'Aktif' olabilir. Lütfen sadece birini seçin.")
                    else:
                        try:
                            sonuc = _sync_donemler(temiz_liste)
                            if sonuc.get("ok"):
                                _fetch_donemler.clear()
                                _fetch_settings.clear()
                                st.session_state.donem_basarili = True
                                st.session_state.exp_donem_acik = True
                                st.rerun()
                            else:
                                st.error(f"Güncellenemedi: {sonuc.get('error', 'Bilinmeyen hata')}")
                        except Exception as e:
                            st.error(f"Güncellenemedi: {e}")

        if "exp_sorular_acik" not in st.session_state:
            st.session_state.exp_sorular_acik = False
        with st.expander("📝 Form Soruları (Ek Sorular) Yönetimi", expanded=st.session_state.exp_sorular_acik):
            st.caption(
                "Buradan eklediğiniz sorular, sabit alanların (T.C., aile bilgileri vb.) ALTINA, "
                "'Ek Sorular' başlığı altında forma otomatik eklenir. Sadece 'Aktif' işaretli sorular "
                "öğrenciye gösterilir. **Seçenekler** sütununu yalnızca Tip='Seçenekli' ise, "
                "seçenekleri virgülle ayırarak doldurun (örn: Evet, Hayır, Bilmiyorum)."
            )
            if st.session_state.get("sorular_basarili"):
                st.success("✅ Sorular güncellendi.")
                st.session_state.sorular_basarili = False
            try:
                sorular_sonuc = _fetch_sorular()
            except Exception as e:
                sorular_sonuc = {"ok": False}
                st.error(f"Sorular alınamadı: {e}")

            if sorular_sonuc.get("ok"):
                sorular_liste = sorular_sonuc.get("sorular", [])
                sorular_df = pd.DataFrame(sorular_liste) if sorular_liste else pd.DataFrame(
                    [{"id": "", "soru": "", "tip": "Kısa Metin", "secenekler": "", "aktif": "Pasif", "sira": 0}]
                )
                for kol in ["id", "soru", "tip", "secenekler", "aktif", "sira"]:
                    if kol not in sorular_df.columns:
                        sorular_df[kol] = ""
                sorular_df = sorular_df[["soru", "tip", "secenekler", "aktif", "sira", "id"]]

                edited_sorular_df = st.data_editor(
                    sorular_df,
                    use_container_width=True,
                    hide_index=True,
                    num_rows="dynamic",
                    column_config={
                        "soru": st.column_config.TextColumn("Soru Metni", required=True, width="large"),
                        "tip": st.column_config.SelectboxColumn(
                            "Tip", options=["Kısa Metin", "Uzun Metin", "Sayı", "Tarih", "Seçenekli", "Evet/Hayır"],
                            required=True,
                        ),
                        "secenekler": st.column_config.TextColumn("Seçenekler (virgülle)"),
                        "aktif": st.column_config.SelectboxColumn("Aktif", options=["Aktif", "Pasif"], required=True),
                        "sira": st.column_config.NumberColumn("Sıra", min_value=0, step=1),
                        "id": st.column_config.TextColumn("ID (dokunmayın)", disabled=True),
                    },
                    key="sorular_editor",
                )

                if st.button("💾 Soruları Kaydet", type="primary"):
                    temiz_liste = [
                        {
                            "id": str(r["id"]).strip() or None,
                            "soru": str(r["soru"]).strip(),
                            "tip": r["tip"],
                            "secenekler": str(r["secenekler"]).strip(),
                            "aktif": r["aktif"],
                        }
                        for _, r in edited_sorular_df.iterrows()
                        if str(r["soru"]).strip()
                    ]
                    try:
                        sonuc = _sync_sorular(temiz_liste)
                        if sonuc.get("ok"):
                            _fetch_sorular.clear()
                            _fetch_sorular_aktif.clear()
                            st.session_state.sorular_basarili = True
                            st.session_state.exp_sorular_acik = True
                            st.rerun()
                        else:
                            st.error(f"Güncellenemedi: {sonuc.get('error', 'Bilinmeyen hata')}")
                    except Exception as e:
                        st.error(f"Güncellenemedi: {e}")

        with st.spinner("Veriler alınıyor..."):
            try:
                data = _fetch_all_rows()
            except Exception as e:
                st.error(f"Veriler alınamadı: {e}")
                st.stop()

        rows = data.get("rows", [])
        if not rows:
            st.info("Henüz başvuru yok.")
        else:
            df_tum = pd.DataFrame(rows)
            for col in ["Babanın Aylık Geliri (TL)", "Annenin Aylık Geliri (TL)", "Okumakta Olan Kardeş Sayısı"]:
                if col in df_tum.columns:
                    df_tum[col] = pd.to_numeric(df_tum[col], errors="coerce")
            if "Zaman Damgası" in df_tum.columns:
                df_tum["Zaman Damgası"] = pd.to_datetime(
                    df_tum["Zaman Damgası"], errors="coerce"
                ).dt.strftime("%d/%m/%Y %H:%M").fillna(df_tum["Zaman Damgası"])
            for tarih_kolonu in ["Öğrenci Doğum Tarihi", "Baba Doğum Tarihi", "Anne Doğum Tarihi"]:
                if tarih_kolonu in df_tum.columns:
                    donusmus = pd.to_datetime(df_tum[tarih_kolonu], errors="coerce").dt.strftime("%d/%m/%Y")
                    df_tum[tarih_kolonu] = donusmus.fillna(df_tum[tarih_kolonu])

            if "Dönem" in df_tum.columns:
                donemler = ["Tümü"] + sorted(df_tum["Dönem"].dropna().unique().tolist(), reverse=True)
                secilen_donem = st.selectbox("📅 Dönem Filtrele", donemler)
                df = df_tum if secilen_donem == "Tümü" else df_tum[df_tum["Dönem"] == secilen_donem]
            else:
                df = df_tum

            if "Form Tipi" in df.columns:
                form_tipleri = ["Tümü"] + sorted(df["Form Tipi"].dropna().unique().tolist())
                secilen_form_tipi = st.selectbox("🧾 Form Tipi Filtrele", form_tipleri)
                if secilen_form_tipi != "Tümü":
                    df = df[df["Form Tipi"] == secilen_form_tipi].copy()
                    # Bu form tipinde tamamen boş kalan sütunları gizle (temiz görünüm/indirme için)
                    df = df.dropna(axis=1, how="all")
                    df = df.loc[:, ~(df.astype(str).apply(lambda s: s.str.strip()) == "").all()]

            tab_genel, tab_liste, tab_evrak = st.tabs(
                ["📊 Genel Bakış", "📋 Başvuru Listesi", "📁 Evrak Klasörleri"]
            )

            # -------------------- GENEL BAKIŞ --------------------
            with tab_genel:
                incelenen = int((df.get("İncelendi", pd.Series(dtype=str)) == "Evet").sum()) if "İncelendi" in df.columns else 0
                m1, m2, m3, m4 = st.columns(4)
                m1.metric("Toplam Başvuru", len(df))
                m2.metric("İncelenen", incelenen)
                m3.metric("Bekleyen", len(df) - incelenen)
                if "Babanın Aylık Geliri (TL)" in df.columns and "Annenin Aylık Geliri (TL)" in df.columns:
                    ort_gelir = (df["Babanın Aylık Geliri (TL)"].fillna(0) + df["Annenin Aylık Geliri (TL)"].fillna(0)).mean()
                    m4.metric("Ort. Aile Geliri (TL)", f"{ort_gelir:,.0f}".replace(",", "."))

                st.divider()
                if "analiz_goster" not in st.session_state:
                    st.session_state.analiz_goster = False

                if not st.session_state.analiz_goster:
                    if st.button("📊 Analizleri Göster (Grafikler)", use_container_width=True):
                        st.session_state.analiz_goster = True
                        st.rerun()
                else:
                    if st.button("🔽 Analizleri Gizle", use_container_width=True):
                        st.session_state.analiz_goster = False
                        st.rerun()

                    donem_sayilari_df = None
                    if "Dönem" in df_tum.columns and df_tum["Dönem"].nunique() > 1:
                        donem_sayilari_df = df_tum["Dönem"].value_counts().reset_index()

                    grafikler = _grafikleri_olustur(df, donem_sayilari_df)

                    g1, g2 = st.columns(2)
                    with g1:
                        if "cinsiyet" in grafikler:
                            st.plotly_chart(grafikler["cinsiyet"], use_container_width=True)
                    with g2:
                        if "sinif" in grafikler:
                            st.plotly_chart(grafikler["sinif"], use_container_width=True)

                    g3, g4 = st.columns(2)
                    with g3:
                        if "bolum" in grafikler:
                            st.plotly_chart(grafikler["bolum"], use_container_width=True)
                    with g4:
                        if "il" in grafikler:
                            st.plotly_chart(grafikler["il"], use_container_width=True)

                    g5, g6 = st.columns(2)
                    with g5:
                        if "medeni_durum" in grafikler:
                            st.plotly_chart(grafikler["medeni_durum"], use_container_width=True)
                    with g6:
                        if "meslek" in grafikler:
                            st.plotly_chart(grafikler["meslek"], use_container_width=True)

                    g7, g8 = st.columns(2)
                    with g7:
                        if "radar" in grafikler:
                            st.plotly_chart(grafikler["radar"], use_container_width=True)
                            st.caption(
                                "Not: Eksenler farklı birimlerdeki değerleri (TL, kişi sayısı, %) "
                                "karşılaştırılabilir kılmak için ölçeklendirilmiştir; mutlak değerler için "
                                "Başvuru Listesi sekmesindeki tabloya bakın."
                            )
                    with g8:
                        if "gelir_histogram" in grafikler:
                            st.plotly_chart(grafikler["gelir_histogram"], use_container_width=True)

                    if "donem_karsilastirma" in grafikler:
                        st.plotly_chart(grafikler["donem_karsilastirma"], use_container_width=True)

            # -------------------- BAŞVURU LİSTESİ (İncelendi işaretleme) --------------------
            with tab_liste:
                st.caption("İncelendi kutucuğunu işaretleyip **Değişiklikleri Kaydet**'e basın.")
                display_cols = [c for c in df.columns if not c.startswith("Belge:")]
                edit_df = df[display_cols].copy()
                if "İncelendi" in edit_df.columns:
                    edit_df["İncelendi"] = edit_df["İncelendi"].apply(lambda v: str(v).strip() == "Evet")

                edited = st.data_editor(
                    edit_df,
                    use_container_width=True,
                    hide_index=True,
                    disabled=[c for c in display_cols if c != "İncelendi"],
                    column_config={
                        "İncelendi": st.column_config.CheckboxColumn("İncelendi", default=False),
                    },
                    key="editor",
                )

                if st.button("💾 Değişiklikleri Kaydet", type="primary"):
                    degisen = 0
                    for i in range(len(edit_df)):
                        eski = edit_df.iloc[i]["İncelendi"]
                        yeni = edited.iloc[i]["İncelendi"]
                        if eski != yeni:
                            tc_no = edit_df.iloc[i].get("T.C. Kimlik No", "")
                            try:
                                _update_review_status(tc_no, "Evet" if yeni else "Hayır")
                                degisen += 1
                            except Exception as e:
                                st.error(f"{tc_no} güncellenemedi: {e}")
                    if degisen:
                        st.success(f"{degisen} kayıt güncellendi.")
                        _fetch_all_rows.clear()
                        st.rerun()
                    else:
                        st.info("Değişiklik bulunamadı.")

                buffer = BytesIO()
                with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                    veri_df = df[display_cols].copy()
                    veri_df.to_excel(writer, index=False, sheet_name="Basvurular")

                    # -------- Özet Pivot sayfası (statik, hazır özet tablolar) --------
                    ozet_sheet_yazildi = False
                    if "Dönem" in veri_df.columns and "Cinsiyet" in veri_df.columns:
                        pivot1 = pd.pivot_table(
                            veri_df, index="Dönem", columns="Cinsiyet",
                            values="T.C. Kimlik No", aggfunc="count", fill_value=0,
                        )
                        pivot1.to_excel(writer, sheet_name="Özet Pivot", startrow=1)
                        ozet_sheet_yazildi = True
                    if "Bölüm" in veri_df.columns:
                        pivot2 = (
                            veri_df["Bölüm"].value_counts().rename("Başvuru Sayısı").to_frame()
                        )
                        baslangic_satiri = len(pivot1) + 5 if ozet_sheet_yazildi else 1
                        pivot2.to_excel(writer, sheet_name="Özet Pivot", startrow=baslangic_satiri)
                        ozet_sheet_yazildi = True

                    ws = writer.sheets["Basvurular"]
                    from openpyxl.worksheet.table import Table, TableStyleInfo
                    from openpyxl.utils import get_column_letter

                    n_satir, n_sutun = veri_df.shape
                    if n_satir > 0 and n_sutun > 0:
                        son_hucre = f"{get_column_letter(n_sutun)}{n_satir + 1}"
                        tablo = Table(displayName="BasvuruTablosu", ref=f"A1:{son_hucre}")
                        tablo.tableStyleInfo = TableStyleInfo(
                            name="TableStyleMedium9", showRowStripes=True,
                            showFirstColumn=False, showLastColumn=False,
                        )
                        ws.add_table(tablo)

                        # Sütun genişliklerini kabaca içeriğe göre ayarla
                        for i, kolon in enumerate(veri_df.columns, start=1):
                            uzunluk = max(len(str(kolon)), veri_df[kolon].astype(str).str.len().max() if n_satir else 0)
                            ws.column_dimensions[get_column_letter(i)].width = min(max(uzunluk + 2, 10), 40)

                        # Gelir / sayısal sütunlara binlik ayraçlı format ver
                        for kolon_adi in ["Babanın Aylık Geliri (TL)", "Annenin Aylık Geliri (TL)",
                                          "Okumakta Olan Kardeş Sayısı"]:
                            if kolon_adi in veri_df.columns:
                                col_idx = veri_df.columns.get_loc(kolon_adi) + 1
                                harf = get_column_letter(col_idx)
                                for satir in range(2, n_satir + 2):
                                    ws[f"{harf}{satir}"].number_format = "#,##0"

                st.download_button(
                    "📥 Excel Olarak İndir (Tablo + Özet Pivot)",
                    data=buffer.getvalue(),
                    file_name=f"basvurular_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
                st.caption(
                    "İndirilen dosyada 'Basvurular' sayfası filtrelenebilir bir Excel tablosu, "
                    "'Özet Pivot' sayfası ise dönem/cinsiyet/bölüm bazlı hazır özet tablolar içerir. "
                    "Kendi dinamik pivot tablonuzu oluşturmak isterseniz: tabloyu seçip "
                    "Excel'de Ekle → PivotTable kullanabilirsiniz."
                )

                # -------------------- KAYIT SİLME (kırmızı buton + onay) --------------------
                st.divider()
                st.markdown("#### 🗑️ Kayıt Silme")
                st.caption(
                    "Silmek istediğiniz kaydın yanındaki kırmızı butona basın — kayıt listeden kaldırılır "
                    "(henüz silinmez). Ardından **Silme İşlemini Uygula**'ya basıp onaylayınca kalıcı olarak silinir."
                )

                if "silinecek_kayitlar" not in st.session_state:
                    st.session_state.silinecek_kayitlar = set()
                if "silme_basarili_sayisi" in st.session_state:
                    n = st.session_state.pop("silme_basarili_sayisi")
                    if n:
                        st.success(f"{n} kayıt silindi.")

                gorunur_satirlar = [
                    (
                        str(r.get("T.C. Kimlik No", "")),
                        str(r.get("Dönem", "")),
                        str(r.get("Öğrenci Adı Soyadı", "—")),
                        str(r.get("Bölüm", "—")),
                    )
                    for _, r in df.iterrows()
                    if (str(r.get("T.C. Kimlik No", "")), str(r.get("Dönem", ""))) not in st.session_state.silinecek_kayitlar
                ]

                if not gorunur_satirlar:
                    st.info("Gösterilecek kayıt yok.")
                else:
                    for tc_no, donem_deger, ad, bolum in gorunur_satirlar:
                        c1, c2, c3, c4 = st.columns([3, 2, 2, 1])
                        c1.write(ad)
                        c2.write(bolum)
                        c3.write(donem_deger)
                        if c4.button("🗑️", key=f"kaldir_{tc_no}_{donem_deger}", help="Listeden kaldır"):
                            st.session_state.silinecek_kayitlar.add((tc_no, donem_deger))
                            st.rerun()

                if st.session_state.silinecek_kayitlar:
                    st.warning(f"{len(st.session_state.silinecek_kayitlar)} kayıt silinmek üzere işaretlendi.")
                    col_uygula, col_vazgec = st.columns(2)
                    with col_uygula:
                        if st.button("🗑️ Silme İşlemini Uygula", type="primary", use_container_width=True):
                            st.session_state["silme_onay_bekliyor"] = True
                    with col_vazgec:
                        if st.button("İşaretleri Temizle", use_container_width=True):
                            st.session_state.silinecek_kayitlar = set()
                            st.rerun()

                if st.session_state.get("silme_onay_bekliyor"):
                    _silme_onay_dialogu()

            # -------------------- EVRAK KLASÖRLERİ --------------------
            with tab_evrak:
                if "Drive Klasör Linki" in df.columns:
                    for _, row in df.iterrows():
                        durum = "✅ İncelendi" if str(row.get("İncelendi", "")).strip() == "Evet" else "⏳ Bekliyor"
                        st.markdown(
                            f"- **{row.get('Öğrenci Adı Soyadı', '—')}** ({durum}) → "
                            f"[Drive Klasörünü Aç]({row.get('Drive Klasör Linki', '#')})"
                        )

        if st.button("Çıkış Yap"):
            st.session_state.admin_authed = False
            st.rerun()
